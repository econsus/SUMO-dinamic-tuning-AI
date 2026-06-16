from typing import List, Dict
from openpyxl import load_workbook

data_records: List[List[Dict[str, int]]] = []
key_names = ["SL", "SPT", "UL", "UPT", "OS", "OU"]
data_path = "CCTV Data Remastered.xlsx"
wb = load_workbook(data_path, data_only = True)

def add_records(data: Dict[str, int], index: int) -> None:
    data_records[index].append(data)

def add_record_row(data: List[List]) -> None:
    data_records.append(data)

def import_records() -> None:
    ws = wb.active
    for row in ws['I3':'N30']:
        values = []
        row_has_data = False

        for index, cell in enumerate(row):
            value = None
            if cell.value is not None and cell.value != '':
                try:
                    value = int(cell.value)
                    row_has_data = True
                except (ValueError, TypeError):
                    value = None

            values.append({key_names[index]: value})

        if not row_has_data:
            break

        data_records.append(values)

